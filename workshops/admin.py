from django.contrib import admin
from django.http import HttpResponse
from django.utils.html import format_html
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
from .models import Workshop, Registration, Payment, School


class RegistrationInline(admin.TabularInline):
    """Inline display of registrations in Workshop admin"""
    model = Registration
    extra = 0
    readonly_fields = ['registration_number', 'student_name', 'grade', 'email', 'payment_status', 'registered_at']
    can_delete = False
    
    def has_add_permission(self, request, obj=None):
        return False


class PaymentInline(admin.StackedInline):
    """Inline display of payment in Registration admin"""
    model = Payment
    extra = 0
    readonly_fields = ['transaction_id', 'amount', 'payment_status', 'initiated_at', 'completed_at']
    can_delete = False
    
    def has_add_permission(self, request, obj=None):
        return False

@admin.register(School)
class SchoolAdmin(admin.ModelAdmin):
    """Admin interface for School model"""
    list_display = ['name', 'is_active', 'created_at']
    list_filter = ['is_active', 'created_at']
    search_fields = ['name']
    readonly_fields = ['created_at', 'updated_at']
@admin.register(Workshop)
class WorkshopAdmin(admin.ModelAdmin):
    """Admin interface for Workshop model"""
    
    list_display = ['name', 'workshop_date', 'fee_display', 'capacity_display', 'is_active', 'created_at']
    list_filter = ['is_active', 'fee', 'created_at']
    search_fields = ['name', 'description', 'venue']
    readonly_fields = ['created_at', 'updated_at', 'current_registrations', 'available_slots']
    
    fieldsets = (
        ('Workshop Information', {
            'fields': ('name', 'description', 'organizer')
        }),
        ('Schedule & Venue', {
            'fields': ('workshop_date', 'time', 'duration', 'venue')
        }),
        ('Registration Details', {
            'fields': ('fee', 'capacity', 'is_active')
        }),
        ('Statistics', {
            'fields': ('current_registrations', 'available_slots', 'created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    
    inlines = [RegistrationInline]
    
    actions = ['export_workshop_registrations']
    
    def fee_display(self, obj):
        """Display fee with FREE tag for free workshops"""
        if obj.is_free:
            return format_html('<span style="color: green; font-weight: bold;">FREE</span>')
        return f"৳{obj.fee}"
    fee_display.short_description = "Fee"
    
    def capacity_display(self, obj):
        """Display capacity with color coding"""
        current = obj.current_registrations
        total = obj.capacity
        percent = (current / total * 100) if total > 0 else 0
        
        if percent >= 90:
            color = 'red'
        elif percent >= 70:
            color = 'orange'
        else:
            color = 'green'
        
        return format_html(
            '<span style="color: {};">{}/{} ({}%)</span>',
            color, current, total, f"{percent:.0f}"
        )
    capacity_display.short_description = "Capacity"
    
    def export_workshop_registrations(self, request, queryset):
        """Export all registrations for selected workshops to Excel"""
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Workshop Registrations"
        
        # Define headers
        headers = [
            'Registration Number', 'Workshop', 'Student Name', 'Grade', 
            'School', 'Contact', 'Email', 'Payment Status', 
            'Fee', 'Registered Date'
        ]
        
        # Style for headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Write data
        row_num = 2
        for workshop in queryset:
            for reg in workshop.registrations.all():
                ws.cell(row=row_num, column=1, value=reg.registration_number)
                ws.cell(row=row_num, column=2, value=workshop.name)
                ws.cell(row=row_num, column=3, value=reg.student_name)
                ws.cell(row=row_num, column=4, value=reg.grade)
                ws.cell(row=row_num, column=5, value=reg.school.name if reg.school else reg.school_name)
                ws.cell(row=row_num, column=6, value=reg.contact_number)
                ws.cell(row=row_num, column=7, value=reg.email)
                ws.cell(row=row_num, column=8, value=reg.get_payment_status_display())
                ws.cell(row=row_num, column=9, value=float(workshop.fee))
                ws.cell(row=row_num, column=10, value=reg.registered_at.strftime('%Y-%m-%d %H:%M:%S'))
                row_num += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=workshop_registrations.xlsx'
        
        return response
    
    export_workshop_registrations.short_description = "Export registrations to Excel"


@admin.register(Registration)
class RegistrationAdmin(admin.ModelAdmin):
    """Admin interface for Registration model"""
    
    list_display = [
        'registration_number', 'student_name', 'grade', 'workshop', 
        'payment_status_display', 'registered_at'
    ]
    list_filter = ['payment_status', 'grade', 'workshop', 'school', 'registered_at']
    search_fields = ['registration_number', 'student_name', 'email', 'school__name', 'school_name']
    readonly_fields = ['registration_number', 'registered_at', 'updated_at']
    
    fieldsets = (
        ('Registration Details', {
            'fields': ('registration_number', 'workshop', 'payment_status')
        }),
        ('Student Information', {
            'fields': ('student_name', 'grade', 'school', 'contact_number', 'email')
        }),
        ('Timestamps', {
            'fields': ('registered_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    
    inlines = [PaymentInline]
    
    actions = ['export_registrations', 'mark_as_completed']
    
    def payment_status_display(self, obj):
        """Display payment status with color coding"""
        colors = {
            'pending': 'orange',
            'completed': 'green',
            'failed': 'red',
            'cancelled': 'gray',
            'free': 'blue',
        }
        color = colors.get(obj.payment_status, 'black')
        return format_html(
            '<span style="color: {}; font-weight: bold;">{}</span>',
            color, obj.get_payment_status_display()
        )
    payment_status_display.short_description = "Payment Status"
    
    def export_registrations(self, request, queryset):
        """Export selected registrations to Excel"""
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Registrations"
        
        # Define headers
        headers = [
            'Registration Number', 'Workshop', 'Workshop Date', 'Student Name', 
            'Grade', 'School', 'Contact', 'Email', 'Payment Status', 
            'Fee', 'Registered Date'
        ]
        
        # Style for headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Write data
        for row_num, reg in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1, value=reg.registration_number)
            ws.cell(row=row_num, column=2, value=reg.workshop.name)
            ws.cell(row=row_num, column=3, value=reg.workshop.workshop_date)
            ws.cell(row=row_num, column=4, value=reg.student_name)
            ws.cell(row=row_num, column=5, value=reg.grade)
            ws.cell(row=row_num, column=6, value=reg.school.name if reg.school else reg.school_name)
            ws.cell(row=row_num, column=7, value=reg.contact_number)
            ws.cell(row=row_num, column=8, value=reg.email)
            ws.cell(row=row_num, column=9, value=reg.get_payment_status_display())
            ws.cell(row=row_num, column=10, value=float(reg.workshop.fee))
            ws.cell(row=row_num, column=11, value=reg.registered_at.strftime('%Y-%m-%d %H:%M:%S'))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=registrations.xlsx'
        
        return response
    
    export_registrations.short_description = "Export selected to Excel"
    
    def mark_as_completed(self, request, queryset):
        """Mark selected registrations as completed"""
        updated = queryset.update(payment_status='completed')
        self.message_user(request, f"{updated} registrations marked as completed.")
    
    mark_as_completed.short_description = "Mark as completed"


@admin.register(Payment)
class PaymentAdmin(admin.ModelAdmin):
    """Admin interface for Payment model"""
    
    list_display = [
        'transaction_id', 'registration', 'amount_display', 
        'payment_status', 'payment_method', 'initiated_at'
    ]
    list_filter = ['payment_status', 'payment_method', 'initiated_at']
    search_fields = ['transaction_id', 'registration__registration_number', 'registration__student_name']
    readonly_fields = ['transaction_id', 'initiated_at', 'completed_at', 'sslcommerz_data']
    
    fieldsets = (
        ('Transaction Details', {
            'fields': ('transaction_id', 'registration', 'amount', 'currency')
        }),
        ('Payment Information', {
            'fields': ('payment_status', 'payment_method')
        }),
        ('Timestamps', {
            'fields': ('initiated_at', 'completed_at')
        }),
        ('SSLCommerz Data', {
            'fields': ('sslcommerz_data',),
            'classes': ('collapse',)
        }),
    )
    
    actions = ['export_payments']
    
    def amount_display(self, obj):
        """Display amount with currency"""
        return f"৳{obj.amount}"
    amount_display.short_description = "Amount"
    
    def export_payments(self, request, queryset):
        """Export selected payments to Excel"""
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Payments"
        
        # Define headers
        headers = [
            'Transaction ID', 'Registration Number', 'Student Name', 'School',
            'Workshop', 'Amount', 'Currency', 'Payment Status', 
            'Payment Method', 'Initiated At', 'Completed At'
        ]
        
        # Style for headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Write data
        for row_num, payment in enumerate(queryset, 2):
            reg = payment.registration
            ws.cell(row=row_num, column=1, value=payment.transaction_id)
            ws.cell(row=row_num, column=2, value=reg.registration_number)
            ws.cell(row=row_num, column=3, value=reg.student_name)
            ws.cell(row=row_num, column=4, value=reg.school.name if reg.school else reg.school_name)
            ws.cell(row=row_num, column=5, value=reg.workshop.name)
            ws.cell(row=row_num, column=6, value=float(payment.amount))
            ws.cell(row=row_num, column=7, value=payment.currency)
            ws.cell(row=row_num, column=8, value=payment.payment_status)
            ws.cell(row=row_num, column=9, value=payment.get_payment_method_display())
            ws.cell(row=row_num, column=10, value=payment.initiated_at.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row_num, column=11, value=payment.completed_at.strftime('%Y-%m-%d %H:%M:%S') if payment.completed_at else 'N/A')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=payments.xlsx'
        
        return response
    
    export_payments.short_description = "Export selected to Excel"
